import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import argparse
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import openpyxl
from openpyxl.styles import Font
from time import sleep
import tkinter as tk


def excelUpdate():
    # Load the Excel workbook
    wb = openpyxl.load_workbook('entities_list.xlsx')
    sheet_name = 'Crypto Companies'  # Replace with your sheet name
    sheet = wb[sheet_name]

    # Task 1: Delete rows that are completely empty
    rows_to_delete = []
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
        if all(cell.value is None for cell in row):
            rows_to_delete.append(row[0].row)

    for row_index in reversed(rows_to_delete):
        sheet.delete_rows(row_index)

    # Task 2: Bold every word in the first row
    for cell in sheet[1]:
        if cell.value:
            cell.font = Font(bold=True)

    # Task 3: Make every other piece of text uniform (same font, size, color, etc.)
    # Example of setting uniform style for all cells
    for row in sheet.iter_rows(min_row=2, min_col=1):
        for cell in row:
            cell.font = Font(name='Arial', size=12, color='000000')  # Adjust as needed

    # Save the updated workbook
    print("this works")
    wb.save('updated_entities_list.xlsx')

def keywordSearch(): 
    df = pd.read_csv('test.csv')
    print(df.columns)

    yes_keywords = ['sustainability', 'energy', 'procurement', 'rec', 'environmental']
    maybe_keywords = ['manager', 'director', 'head', 'chief']

    def classify_title(title):
        if pd.isna(title):
            return 'No'
        title_lower = str(title).lower()
        if any(keyword in title_lower for keyword in yes_keywords):
            return 'Yes'
        elif any(keyword in title_lower for keyword in maybe_keywords):
            return 'Maybe'
        else:
            return 'No'

    df['I'] = df['Title'].apply(classify_title)
    df.to_csv('test.csv', index=False)
    print("Updated file with REC Responsibility in column 'I'.")

def apolloAutomation():
    df = pd.read_csv('test.csv')
    
    # Set up the WebDriver without headless mode to see the browser interactions
    driver = webdriver.Chrome()  # This will open the browser window so you can see the interactions

    def login():
        driver.get('https://app.apollo.io/#/login?redirectTo=https%3A%2F%2Fapp.apollo.io%2F%23%2F')
        print("Navigated to login page")

        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.NAME, "email")))
        print("Email input found")
        driver.find_element(By.NAME, 'email').send_keys('austin25@illinois.edu')
        driver.find_element(By.NAME, 'password').send_keys('ourmission3888')
        driver.find_element(By.NAME, 'password').send_keys(Keys.RETURN)
        print("Login credentials submitted")
        return True
    
    def xpath_is_there(xpath):
        sleep(3)
        elements = driver.find_elements(By.XPATH, xpath)
        if elements:
            print("Xpath is found")
        else: 
            print("Xpath not found")

    def is_element_present(xpath):
        try:
            elements = driver.find_elements(By.XPATH, xpath)
            return len(elements) > 0
        except NoSuchElementException:
            return False
    
    def is_css_selector_present(css_selector):
        try:
            elements = driver.find_elements(By.CSS_SELECTOR, css_selector)
            print("yes this CSS is present")
            return len(elements) > 0
        except NoSuchElementException:
            print("na its not present")
            return False
    
    def wait_for_element(driver, xpath, timeout=10):
        try:
            element = WebDriverWait(driver, timeout).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            print(f"Element is found: {xpath}")
            return element
        except TimeoutException:
            print(f"Element not found: {xpath}")
            return None


    def search_and_retrieve(first_name, last_name, company, job_title):
        try:
            
            search_bar = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//input[@placeholder='Search Apollo']")))
            print("Search bar found")
            search_bar.clear()
            search_bar.send_keys(Keys.CONTROL + "a")
            search_bar.send_keys(Keys.DELETE)  
            # preparing quqery
            search_query = f"{first_name} {last_name} {company} {job_title}"
            search_bar.send_keys(search_query)
            search_bar.send_keys(Keys.RETURN)

            # Wait for the top search result
            top_result_selector = "div.zp_eNBrc.zp_gD9FN.zp_uYsJY"
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, top_result_selector)))
            print("Search results found")

            # Click the top search result
            time.sleep(3)
            top_result = driver.find_element(By.CSS_SELECTOR, top_result_selector)
            top_result.click()

            # Find the LinkedIn profile URL
            linkedin_url_xpath =  '//*[@id="main-app"]/div[2]/div/div/div[2]/div[2]/div/div/div[1]/div/div[2]/div[2]/div/a'
            wait = WebDriverWait(driver, 10)
            xpath_is_there(linkedin_url_xpath)
            linkedin_element = wait.until(EC.presence_of_element_located((By.XPATH, linkedin_url_xpath)))

            #Checking to make sure its the same webpage
            url_before = driver.current_url
            print("URL Before is: ", url_before)

            xpath_is_there(linkedin_url_xpath)
            linkedin_element.click()
            time.sleep(5)
            handles = driver.window_handles
            driver.switch_to.window(handles[1])
            their_linked_in = driver.current_url
            print("Current URL:", their_linked_in)
            time.sleep(3)
            driver.close()

            # Find the Email Address
            driver.switch_to.window(handles[0])
            time.sleep(10)

            email_xpath = '(//a[@class="zp_p4M34 zp_q6Sul zp_dAPkM zp_Iu6Pf"])[6]'
            wait = WebDriverWait(driver, 10)
            xpath_is_there(email_xpath)
            email_element = wait.until(EC.presence_of_element_located((By.XPATH, linkedin_url_xpath)))

            
            # Locate the element by its class name
            
            
            xpath = '//*[@id="ly1ogihk"]/div[2]/div/div[1]/div[2]/div/div/div[1]/div/div[2]/div/div/div[2]/div[1]/span/button/i'
    
            # Find all elements matching the XPath
            element = wait.until(EC.presence_of_all_elements_located((By.XPATH, xpath)))
            xpath_is_there(xpath)
            
            # Iterate through each element and print its text content


            
            return "No"

            
        except TimeoutException:
            print("No search results found for query:", search_query)
            return "Not on Apollo"

# Example usage:
# driver = webdriver.Chrome()
# login()  # Assuming the login function is defined elsewhere and logs in successfully
# result = search_and_retrieve("John", "Doe", "Example Company", "Software Engineer")
# print(result)

    login()
    for index, row in df.iterrows():
        email_or_none = search_and_retrieve(row['First Name'], row['Last Name'], row['Company'], row['Title'])
        df.loc[index, 'Email'] = email_or_none
        print("My output into excecl is: ", email_or_none)

    df.to_csv('ceba_with_email.csv', index=False)
    print("It worked")
    
    # driver.quit()

def main():
    parser = argparse.ArgumentParser(description="Run specified function from the terminal")
    parser.add_argument('function_name', type=str, help="The name of the function to run")
    args = parser.parse_args()

    if args.function_name == "keywordSearch":
        keywordSearch()
    elif args.function_name == "apolloAutomation":
        apolloAutomation()
    else:
        print(f"Function '{args.function_name}' not found")

if __name__ == "__main__":
    apolloAutomation()
    print("Here we go")
    

